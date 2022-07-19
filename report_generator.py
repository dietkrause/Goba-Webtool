import pandas as pd
import csv
from openpyxl.styles import Border, Side
from logging import warning
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
import os
import sys
#import PIL
import io
import urllib3
from openpyxl.drawing.image import Image




def receive_factor_sql_file(file):     
    date=None
    header=None
    data=[]
    with open(file, 'r') as csvfile:
        csvreader = csv.reader(csvfile)
        d=True
        h=True
        for i in csvreader:
            if d:
                date=i
                d=False
            elif h:
                header=i
                h=False
            else:
                data.append(i)
        products={"FACTORING":[],"SUPLI-CREDIT":[],"INVENTORY-FINANCING":[]}
        first=True
        for row in data:
            if first:
                company=row[0].split("-")[1]
                first=False
            if "factoring" in row[0].lower():
                products["FACTORING"].append(row)
            elif "supli" in row[0].lower() or "suppli" in row[0].lower():
                products["SUPLI-CREDIT"].append(row)
            elif "inventory" in row[0].lower():
                products["INVENTORY-FINANCING"].append(row)


        products_2={"FACTORING":[],"SUPLI-CREDIT":[],"INVENTORY-FINANCING":[]}
        for key in products:
            comp={}
            for row in products[key]:
                nrow=[]
                for i in row[6:]:
                    try:
                        nrow.append(float(i))
                    except ValueError:
                        nrow.append(i)
                if "POs" in row[2]:
                    clasy=" - POs"
                elif "INV" in row[2] and "INVENTORY" not in row[2]:
                    clasy=" - INV"
                elif "STG" in row[2]:
                    clasy=" - INV"
                else:
                    clasy=""
                nrow=[row[5]+clasy]+nrow
                if row[3] not in comp:
                    comp[row[3]]=[]
                    comp[row[3]].append(nrow)
                else:
                    comp[row[3]].append(nrow)
            products_2[key]=comp

        date={'Start_date':date[1],'End_date':date[3]}
        return date,header,products_2,company

def from_xslx_to_csv(file):
    if "$" not in file:
        read_file = pd.read_excel (file)
        f=file.split(".")
        FILE=f[0]+".csv"
        read_file.to_csv (FILE, index = None, header=True)
        return FILE

def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def set_space(ws, cell_range):
    thin = Side(border_style="thin", color="FFFFFF")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

#----------------------------------------------------------------------
def create_from_dir(DIR_FROM:str,DIR_TO:str):
    path="Portfolio_Report"
    path+=".xlsx"
    path=os.path.join(DIR_TO,path)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    workbook.save(path)


    xfile = openpyxl.load_workbook(path,data_only=True)
    sheet = xfile.get_sheet_by_name('Sheet')
    xfile.save(path)


    #------------------ HOUSEKEEPING --------------
    LET="ABCDEFGHIJKLMNOPQRST"
    green = PatternFill(patternType='solid', 
                                fgColor='6B8E23')
    grey = PatternFill(patternType='solid', 
                                fgColor='808080')
    r = 1
    http = urllib3.PoolManager()
    r = http.request('GET', 'https://files.monday.com/use1/email_headers/7979575/thumb/goba_capital_email_header_726751.png')
    image_file = io.BytesIO(r.data)
    img = Image(image_file)
    img.anchor = 'D1'
    sheet.add_image(img)
    sheet.merge_cells('D1:E3')

    ST=4
    set_space(sheet,"A1:K{}".format(ST+2))
    dir=os.listdir(DIR_FROM)
    for f in dir:
        country=f.split(".")[0]
        user_file=os.path.join(DIR_FROM,f)
        try:
            user_file=os.path.abspath(user_file)
            file_name=from_xslx_to_csv(user_file)
            date,header,data,company=receive_factor_sql_file(file_name)

            sheet.column_dimensions["A"].width = 50
            sheet.column_dimensions["B"].width = 50
            sheet.column_dimensions["C"].width = 30
            sheet.column_dimensions["D"].width = 20
            sheet.column_dimensions["E"].width = 20
            sheet.column_dimensions["F"].width = 20
            sheet.column_dimensions["G"].width = 20
            sheet.column_dimensions["H"].width = 20
            sheet.column_dimensions["I"].width = 20
            sheet.column_dimensions["J"].width = 20
            sheet.column_dimensions["K"].width = 40
            sheet['A{}'.format(ST)] = "Outstanding Portfolio Goba Capital {}".format(country)
            sheet['A{}'.format(ST)].font = Font(bold=True)
            ST+=1
            sheet['A{}'.format(ST)]="As of {}".format(date['Start_date'])
            sheet['A{}'.format(ST)].font = Font(bold=True)
            ST+=3
            sheet['A{}'.format(ST)]="Client's name"
            sheet['A{}'.format(ST)].font = Font(bold=True)
            sheet.merge_cells('A{}:B{}'.format(ST-5,ST-5))
            sheet['C{}'.format(ST)]="Total Invoice amount"
            sheet['C{}'.format(ST)].font = Font(bold=True)
            sheet['D{}'.format(ST)]="Current"
            sheet['D{}'.format(ST)].font = Font(bold=True)
            sheet['E{}'.format(ST)]="1 to 30"
            sheet['E{}'.format(ST)].font = Font(bold=True)
            sheet['F{}'.format(ST)]="31 to 60"
            sheet['F{}'.format(ST)].font = Font(bold=True)
            sheet['G{}'.format(ST)]="61 to 90"
            sheet['G{}'.format(ST)].font = Font(bold=True)
            sheet['H{}'.format(ST)]="91 to 120"
            sheet['H{}'.format(ST)].font = Font(bold=True)
            sheet['I{}'.format(ST)]="121 to 150"
            sheet['I{}'.format(ST)].font = Font(bold=True)
            sheet['J{}'.format(ST)]="Over 150"
            sheet['J{}'.format(ST)].font = Font(bold=True)
            sheet['K{}'.format(ST)]="Total advanced amount"
            sheet['K{}'.format(ST)].font = Font(bold=True)


            for let in LET:
                sheet['{}{}'.format(let,ST)].fill = green
                if let.lower()=="k":
                    break
            #-----------------------------------------------
            ST_ABS=ST
            ST_ROW=ST
            ED_ROW=ST_ROW
            TOTAL=[]
            BORDER=ST
            for key in data:
                if data[key]=={}:
                    continue
                ED_ROW+=1
                sheet['A{}'.format(ED_ROW)]=key
                ED_ROW+=1
                ST_ROW=ED_ROW
                FAC=[]
                for client in data[key]: 
                    sheet['A{}'.format(ED_ROW)]=client
                    for rec in data[key][client]:
                        col=1
                        index=0
                        for val in rec:
                            sheet['{}{}'.format(LET[col],ED_ROW)]=val
                            sheet['{}{}'.format(LET[col],ED_ROW)].number_format ='"$"#,##0_);("$"#,##0)'
                            if col!=1:
                                col+=1
                            else:
                                col+=2
                            index+=1
                        sheet['C{}'.format(ED_ROW)] = '= SUM(D{}:J{})'.format(ED_ROW,ED_ROW) 
                        sheet['C{}'.format(ED_ROW)].number_format ='"$"#,##0_);("$"#,##0)' 
                        ED_ROW+=1
                    sheet['B{}'.format(ED_ROW)]="{} TOTAL".format(client)
                    sheet['B{}'.format(ED_ROW)].font = Font(bold=True)
                    #TOTALS FOR EVERY COLUMN
                    for let in LET[2:]:
                        sheet['{}{}'.format(let,ED_ROW)] = '= SUM({}{}:{}{})'.format(let,ST_ROW,let,ED_ROW-1)
                        sheet['{}{}'.format(let,ED_ROW)].font = Font(bold=True)
                        sheet['{}{}'.format(let,ED_ROW)].number_format ='"$"#,##0_);("$"#,##0)' 
                        if let.lower()=="k":
                            break
                    #TOTALS END
                    sheet['C{}'.format(ED_ROW)].font = Font(bold=True)
                    sheet['C{}'.format(ED_ROW)].number_format ='"$"#,##0_);("$"#,##0)' 
                    FAC.append('C{}'.format(ED_ROW)) #------------------------- LOOP OVER THIS TO SUM ROWS
                    sheet.merge_cells('A{}:A{}'.format(ST_ROW,ED_ROW))
                    ED_ROW+=1
                    ST_ROW=ED_ROW
                sheet['A{}'.format(ED_ROW)]="{} TOTAL".format(key)
                sheet['A{}'.format(ED_ROW)].font = Font(bold=True)

                sheet.merge_cells('A{}:B{}'.format(ED_ROW,ED_ROW))
                sheet["A{}".format(ED_ROW)].fill = grey
                T=""
                for i in FAC:
                    T+=i+"+"
                T=T[:-1]
                for let in LET[2:]:
                    T_new=T.replace("C",let)
                    sheet['{}{}'.format(let,ED_ROW)]="={}".format(T_new)
                    sheet['{}{}'.format(let,ED_ROW)].number_format ='"$"#,##0_);("$"#,##0)' 
                    sheet['{}{}'.format(let,ED_ROW)].font = Font(bold=True)
                    sheet['{}{}'.format(let,ED_ROW)].fill = grey
                    if let.lower()=="k":
                        break
                TOTAL.append('C{}'.format(ED_ROW))

            ED_ROW+=1
            T=""
            for i in TOTAL:
                T+=i+"+"
            T=T[:-1]
            sheet['A{}'.format(ED_ROW)]="TOTAL {} OUTSTANDING".format(country)
            sheet['A{}'.format(ED_ROW)].font = Font(bold=True)
            sheet['C{}'.format(ED_ROW)]="={}".format(T)
            sheet['C{}'.format(ED_ROW)].font = Font(bold=True)
            sheet['C{}'.format(ED_ROW)].number_format ='"$"#,##0_);("$"#,##0)' 

            for let in LET[3:]:
                T_new=T
                sheet['{}{}'.format(let,ED_ROW)]="={}".format(T_new.replace("C",let))
                sheet['{}{}'.format(let,ED_ROW)].number_format ='"$"#,##0_);("$"#,##0)' 
                sheet['{}{}'.format(let,ED_ROW)].font = Font(bold=True)
                sheet['{}{}'.format(let,ED_ROW)].fill = grey
                if let.lower()=="k":
                    break

            sheet.merge_cells('A{}:B{}'.format(ED_ROW,ED_ROW))
            for let in LET:
                sheet['{}{}'.format(let,ED_ROW)].fill = green
                if let.lower()=="k":
                    break
            set_border(sheet, 'A{}:K{}'.format(BORDER,ED_ROW)) 
            set_space(sheet,"A{}:K{}".format(ED_ROW,ED_ROW+5))
            ST=ED_ROW+3
        except:
            warning("THE FILE IS NOT VALID, TRY A NEW ONE")
            continue

    dir=os.listdir(DIR_FROM)
    for f in dir:
        if "csv" in f:
            os.remove(os.path.join(DIR_FROM,f))

    xfile.save('{}'.format(path)) ## REVIEW IF THE IDENTATION CORRESPONDS TO THE LOGIC

def receive_factor_sql_file2(file):
    file=from_xslx_to_csv(file)
    date=None
    header=None
    data=[]
    with open(file, 'r') as csvfile:
        csvreader = csv.reader(csvfile)
        d=True
        h=True
        for i in csvreader:
            if d:
                date=i
                d=False
            elif h:
                header=i
                h=False
            else:
                data.append(i)
        date=date[1]
        countries={"CHILE":{},"COLOMBIA":{},"PERU":{},"CENTROAMERICA":{}}
        first=True
        for row in data:
            if row[7].lower()=="prior":
                continue
            if first:
                company=row[1]
                first=False
            r=row[2]
            rate=round(float(row[3].replace("%",""))/100,1)
            ob=float(row[8])
            aa=rate*ob
            if ob==0:
                ob="-"
                aa="-"
            if "CHIL" in r:
                if "ALLOC" in r or "INV" in r or "STG" in r:
                    if countries["CHILE"]=={}:
                        countries["CHILE"]["INVOICES"]={}
                        countries["CHILE"]["INVOICES"][row[7]]=[ob,aa]
                    else:
                        try:
                            countries["CHILE"]["INVOICES"][row[7]]=[ob,aa]
                        except KeyError:
                            countries["CHILE"]["INVOICES"]={}
                            countries["CHILE"]["INVOICES"][row[7]]=[ob,aa]
                else:
                    if countries["CHILE"]=={}:
                        countries["CHILE"]["POs"]={}
                        countries["CHILE"]["POs"][row[7]]=[ob,aa]
                    else:
                        try:
                            countries["CHILE"]["POs"][row[7]]=[ob,aa]
                        except KeyError:
                            countries["CHILE"]["POs"]={}
                            countries["CHILE"]["POs"][row[7]]=[ob,aa]
                            
            elif "PERU" in r:
                if "ALLOC" in r or "INV" in r or "STG" in r:
                    if countries["PERU"]=={}:
                        countries["PERU"]["INVOICES"]={}
                        countries["PERU"]["INVOICES"][row[7]]=[ob,aa]
                    else:
                        try:
                            countries["PERU"]["INVOICES"][row[7]]=[ob,aa]
                        except KeyError:
                            countries["PERU"]["INVOICES"]={}
                            countries["PERU"]["INVOICES"][row[7]]=[ob,aa]
                else:
                    if countries["PERU"]=={}:
                        countries["PERU"]["POs"]={}
                        countries["PERU"]["POs"][row[7]]=[ob,aa]
                    else:
                        try:
                            countries["PERU"]["POs"][row[7]]=[ob,aa]
                        except KeyError:
                            countries["PERU"]["POs"]={}
                            countries["PERU"]["POs"][row[7]]=[ob,aa]

            elif "COL" in r:
                if "ALLOC" in r or "INV" in r or "STG" in r:
                    if countries["COLOMBIA"]=={}:
                        countries["COLOMBIA"]["INVOICES"]={}
                        countries["COLOMBIA"]["INVOICES"][row[7]]=[ob,aa]
                    else:
                        try:
                            countries["COLOMBIA"]["INVOICES"][row[7]]=[ob,aa]
                        except KeyError:
                            countries["COLOMBIA"]["INVOICES"]={}
                            countries["COLOMBIA"]["INVOICES"][row[7]]=[ob,aa]

                else:
                    if countries["COLOMBIA"]=={}:
                        countries["COLOMBIA"]["POs"]={}
                        countries["COLOMBIA"]["POs"][row[7]]=[ob,aa]
                    else:
                        try:
                            countries["COLOMBIA"]["POs"][row[7]]=[ob,aa]
                        except KeyError:
                            countries["COLOMBIA"]["POs"]={}
                            countries["COLOMBIA"]["POs"][row[7]]=[ob,aa]

            else:
                if "ALLOC" in r or "INV" in r or "STG" in r:
                    if countries["CENTROAMERICA"]=={}:
                        countries["CENTROAMERICA"]["INVOICES"]={}
                        countries["CENTROAMERICA"]["INVOICES"][row[7]]=[ob,aa]
                    else:
                        try:
                            countries["CENTROAMERICA"]["INVOICES"][row[7]]=[ob,aa]
                        except KeyError:
                            countries["CENTROAMERICA"]["INVOICES"]={}
                        countries["CENTROAMERICA"]["INVOICES"][row[7]]=[ob,aa]

                else:
                    if countries["CENTROAMERICA"]=={}:
                        countries["CENTROAMERICA"]["POs"]={}
                        countries["CENTROAMERICA"]["POs"][row[7]]=[ob,aa]
                    else:
                        try:
                            countries["CENTROAMERICA"]["POs"][row[7]]=[ob,aa]
                        except KeyError:
                            countries["CENTROAMERICA"]["POs"]={}
                            countries["CENTROAMERICA"]["POs"][row[7]]=[ob,aa]

        return countries,date,company

def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def set_space(ws, cell_range):
    thin = Side(border_style="thin", color="FFFFFF")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

#----------------------- XXXXXXX -----------------
def create_from_dir2(FROM,TO):
    fil=os.listdir(FROM)
    for f in fil:
        print("current file -->", f)
        if "$" not in f and "xlsx" in f:
            file=f

    p=os.path.join(FROM,file)
    countries,date,company=receive_factor_sql_file2(os.path.abspath(p))

    #-------- File creation -------

    path="Minimum_Outstanding"
    path+=".xlsx"
    today_date=date
    path=os.path.join(TO,path)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    workbook.save(path)


    xfile = openpyxl.load_workbook(path,data_only=True)
    sheet = xfile['Sheet']
    xfile.save(path)

    #------- end of file creation -----


    #------------------ HOUSEKEEPING --------------
    LET="ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    LET_L=[]
    for l in LET:
        LET_L.append(l)
    LET_L2=[]
    for i in LET:
        sheet.column_dimensions[i].width = 20
        for j in LET:
            t="{}{}".format(i,j)
            sheet.column_dimensions[t].width = 20
            LET_L2.append(t)
    LET=LET_L+LET_L2
    green = PatternFill(patternType='solid', 
                                fgColor='6B8E23')
    grey = PatternFill(patternType='solid', 
                                fgColor='808080')

    sheet["B1"]="GOBA CAPITAL INC"
    sheet["B1"].font = Font(bold=True)
    sheet.merge_cells('B1:D1')
    sheet["B2"]="{} Minimum Outstanding".format(company)
    sheet["B2"].font = Font(bold=True)
    sheet.merge_cells('B2:D2')
    sheet["B3"]=" {} ".format(today_date)

    xfile.save(path)
    #----- TABLE LOOP --------
    ST_COL=1
    for country in countries:
        ST_ROW=4
        b_top=[ST_ROW,ST_COL]
        sheet["{}{}".format(LET[ST_COL],ST_ROW)]=country
        sheet["{}{}".format(LET[ST_COL],ST_ROW)].font = Font(bold=True)
        sheet.merge_cells('{}{}:{}{}'.format(LET[ST_COL],ST_ROW,LET[ST_COL+2*len(countries[country])],ST_ROW))
        sheet["{}{}".format(LET[ST_COL],ST_ROW)].fill = green
        ST_ROW+=1
        first_date=True
        ST_init=ST_ROW
        for service in countries[country]:
            if first_date:
                sheet["{}{}".format(LET[ST_COL],ST_ROW)].fill = green
                sheet["{}{}".format(LET[ST_COL+1],ST_ROW)]=service
                sheet["{}{}".format(LET[ST_COL+1],ST_ROW)].font = Font(bold=True)
                sheet["{}{}".format(LET[ST_COL],ST_ROW+1)]="date"
                sheet["{}{}".format(LET[ST_COL],ST_ROW+1)].font = Font(bold=True)
                sheet["{}{}".format(LET[ST_COL],ST_ROW+1)].fill = green
                sheet["{}{}".format(LET[ST_COL+1],ST_ROW+1)]="Outstanding Balance"
                sheet["{}{}".format(LET[ST_COL+1],ST_ROW+1)].font = Font(bold=True)
                sheet["{}{}".format(LET[ST_COL+1],ST_ROW+1)].fill = green
                sheet["{}{}".format(LET[ST_COL+2],ST_ROW+1)]="Advance Amount"
                sheet["{}{}".format(LET[ST_COL+2],ST_ROW+1)].font = Font(bold=True)
                sheet["{}{}".format(LET[ST_COL+2],ST_ROW+1)].fill = green
                sheet.merge_cells('{}{}:{}{}'.format(LET[ST_COL+1],ST_ROW,LET[ST_COL+2],ST_ROW))
                sheet["{}{}".format(LET[ST_COL+1],ST_ROW)].fill = green
                first_date=False
                first_service=True
            else:
                sheet["{}{}".format(LET[ST_COL],ST_ROW)]=service
                sheet["{}{}".format(LET[ST_COL],ST_ROW)].font = Font(bold=True)
                sheet["{}{}".format(LET[ST_COL],ST_ROW+1)]="Outstanding Balance"
                sheet["{}{}".format(LET[ST_COL],ST_ROW+1)].font = Font(bold=True)
                sheet["{}{}".format(LET[ST_COL],ST_ROW+1)].fill = green
                sheet["{}{}".format(LET[ST_COL+1],ST_ROW+1)]="Advance Amount"
                sheet["{}{}".format(LET[ST_COL+1],ST_ROW+1)].font = Font(bold=True)
                sheet["{}{}".format(LET[ST_COL+1],ST_ROW+1)].fill = green
                sheet.merge_cells('{}{}:{}{}'.format(LET[ST_COL],ST_ROW,LET[ST_COL+1],ST_ROW))
                sheet["{}{}".format(LET[ST_COL],ST_ROW)].fill = green
            ST_ROW+=1  
            for date in countries[country][service]:
                ST_ROW+=1
                if first_service:
                    sheet["{}{}".format(LET[ST_COL],ST_ROW)]=date
                    sheet["{}{}".format(LET[ST_COL+1],ST_ROW)]=countries[country][service][date][0]
                    sheet["{}{}".format(LET[ST_COL+1],ST_ROW)].number_format ='"$"#,##0_);("$"#,##0)'
                    sheet["{}{}".format(LET[ST_COL+2],ST_ROW)]=countries[country][service][date][1]
                    sheet["{}{}".format(LET[ST_COL+2],ST_ROW)].number_format ='"$"#,##0_);("$"#,##0)'
                else:
                    sheet["{}{}".format(LET[ST_COL],ST_ROW)]=countries[country][service][date][0]
                    sheet["{}{}".format(LET[ST_COL],ST_ROW)].number_format ='"$"#,##0_);("$"#,##0)'
                    sheet["{}{}".format(LET[ST_COL+1],ST_ROW)]=countries[country][service][date][1]
                    sheet["{}{}".format(LET[ST_COL+1],ST_ROW)].number_format ='"$"#,##0_);("$"#,##0)'

            if first_service:
                sheet["{}{}".format(LET[ST_COL+1],ST_ROW+1)]="=AVERAGE({}{}:{}{})".format(LET[ST_COL+1],b_top[0]+3,LET[ST_COL+1],ST_ROW)
                sheet["{}{}".format(LET[ST_COL+1],ST_ROW+1)].number_format ='"$"#,##0_);("$"#,##0)'
                sheet["{}{}".format(LET[ST_COL+1],ST_ROW+1)].font = Font(bold=True)
                ST_COL+=3
                first_service=False
            else:
                sheet["{}{}".format(LET[ST_COL],ST_ROW+1)]="=AVERAGE({}{}:{}{})".format(LET[ST_COL],b_top[0]+3,LET[ST_COL],ST_ROW)
                sheet["{}{}".format(LET[ST_COL],ST_ROW+1)].number_format ='"$"#,##0_);("$"#,##0)'
                sheet["{}{}".format(LET[ST_COL],ST_ROW+1)].font = Font(bold=True)
                ST_COL+=2
            row_bottom=ST_ROW
            ST_ROW=ST_init

        b_bottom=[row_bottom,ST_COL]
        sheet["{}{}".format(LET[b_top[1]],b_bottom[0]+1)]="AVERAGE"
        sheet["{}{}".format(LET[b_top[1]],b_bottom[0]+1)].font = Font(bold=True)
        set_border(sheet,"{}{}:{}{}".format(LET[b_top[1]],b_top[0],LET[b_bottom[1]-1],b_bottom[0]+1))
        ST_COL+=1
        NEW_ST=b_bottom[0]+4

    #------ TABLE LOOP END ----

    sheet.freeze_panes=sheet["A7"]
    xfile.save(path)
